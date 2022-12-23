from django.utils.crypto import get_random_string
from django.forms import BaseForm

from openpyxl import load_workbook
import filetype
import io
import csv

def get_errors_object (form : BaseForm) -> dict:
    data = {}

    for f in form :
        if len(f.errors) > 0 : data[f.name] = f.errors

    return data

def generate_str (len:str = 10) -> str:
    allowed_chars="abcdefghjkmnpqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return get_random_string(len, allowed_chars)

def view_set_to_list_create (view_set) :
    return view_set.as_view({'get': 'list', 'post': 'create'})

def view_set_to_crud (view_set) :
    return view_set.as_view({
        'get': 'retrieve', 
        'put': 'update', 
        'patch': 'partial_update', 
        'delete': 'destroy'
    })

def parse_excel (file : str | io.BytesIO) -> list[dict]:
    wb = load_workbook(file.file)
    ws = wb.active

    rows = list(ws.iter_rows())
    columns = [cell.value.strip() for cell in rows[0]]
    data_rows = rows[1:]
    data = list()

    for row in data_rows:
        data_row = dict()

        for i, cell in enumerate(row) :
            data_row[columns[i]] = cell.value

        data.append(data_row)

    return data

def parse_csv (file : str | io.BytesIO) -> list[dict]:
    reader = csv.DictReader(file)
    return list(reader)
    

def check_filetype (obj, excepted : list = list()) :
    mime = filetype.guess_mime(obj)

    return mime in excepted